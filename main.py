# main.py
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify
from flask_mysqldb import MySQL, MySQLdb
from werkzeug.utils import secure_filename
#import datetime
import pandas as pd
from flask import send_file
import os
import json
import tempfile
from io import BytesIO
import csv
import openpyxl
import requests
from datetime import datetime, date
import random
from io import BytesIO
#import pandas as pd
from flask import send_file, flash
from datetime import datetime, timedelta


app = Flask(__name__)
app.secret_key = "supersecretkey"  

#mysql database configuration
app.config.update(
    MYSQL_HOST='mysql-10505a59-philalushaba11d-51bd.f.aivencloud.com',
    MYSQL_USER='avnadmin',
    MYSQL_PORT= 13152,
    MYSQL_PASSWORD='AVNS_lurH_tAP0IeniHK2r7b',
    MYSQL_DB='lunyati'
)

# Initialize MySQL
mysql = MySQL(app)

# Helper to get a dict cursor
def dict_cursor():
    return mysql.connection.cursor(MySQLdb.cursors.DictCursor)

def recconnect_mysql():
    try:
        mysql.connection.ping(reconnect=True)
        cur = mysql.connection.cursor()
        cur.execute("SET SESSION wait_timeout=300")
        cur.close()
    except Exception:
        pass


# Landing Page
@app.route('/')
def home():
    return render_template('index.html')


# Login Page
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()

        try:
            cur = dict_cursor()
            cur.execute("SELECT * FROM users WHERE username=%s AND password=%s", (username, password))
            user = cur.fetchone()
            cur.close()
        except Exception as e:
            import traceback
            traceback.print_exc()
            flash(f"Database error: {e}", "danger")
            return render_template('login.html')

        if user:
            session["username"] = username
            flash("Login successful!", "success")
            return redirect(url_for("dashboard"))
        else:
            flash("Invalid username or password.", "danger")

    return render_template('login.html')


# admin Register Page
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        full_name = request.form.get("full_name", "").strip()
        phone = request.form.get("phone", "").strip()
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        role = request.form.get("role", "").strip()

        if not all([full_name, phone, username, password, role]):
            flash("All fields are required.", "danger")
            return redirect(url_for("register"))

        try:
            cur = mysql.connection.cursor()
            cur.execute("SELECT id FROM users WHERE username=%s", (username,))
            if cur.fetchone():
                flash("Username already exists.", "danger")
                cur.close()
            else:
                cur.execute(
                    "INSERT INTO users (full_name, phone, username, password, role) VALUES (%s, %s, %s, %s, %s)",
                    (full_name, phone, username, password, role)
                )
                mysql.connection.commit()
                cur.close()
                flash("Registration successful! Please login.", "success")
                return redirect(url_for("login"))
        except Exception as e:
            flash(f"Database error: {e}", "danger")

    return render_template('register.html')


# Dashboard Page
@app.route('/dashboard')
def dashboard():
    if "username" not in session:
        flash("Please log in first.", "danger")
        return redirect(url_for("login"))

    try:
        cur = dict_cursor()
        cur.execute("SELECT * FROM books WHERE status = 'Available' ORDER BY created_at DESC")
        available_books = cur.fetchall()
        cur.execute("SELECT * FROM books WHERE status = 'Rented' ORDER BY due_date ASC")
        rented_books = cur.fetchall()
        cur.close()
    except Exception as e:
        flash(f"Could not load books: {e}", "danger")
        available_books = []
        rented_books = []

    return render_template('dashboard.html', username=session["username"],
                           available_books=available_books, rented_books=rented_books)


# Logout
@app.route('/logout')
def logout():
    session.pop("username", None)
    flash("You have been logged out.", "info")
    return redirect(url_for("home"))


# --- Add book (POST JSON or form) ---
@app.route('/books/add', methods=['POST'])
def add_book():
    data = request.get_json() or request.form
    title = data.get('title')
    author = data.get('author')
    publisher = data.get('publisher')
    year = data.get('year') or None
    category = data.get('category')
    charge_after_due_date = float(data.get('charge_after_due_date') or 0.00)

    if not all([title, category]):
        return jsonify({"ok": False, "msg": "Title and category are required"}), 400

    try:
        cur = dict_cursor()
        # Insert the book without book_id
        cur.execute("""
            INSERT INTO books (title, author, publisher, year, category, status, charge_after_due_date)
            VALUES (%s, %s, %s, %s, %s, 'Available', %s)
        """, (title, author, publisher, year, category, charge_after_due_date))

        # Get the last inserted auto-incremented ID
        book_id_number = cur.lastrowid
        # Generate the book_id with the prefix
        book_id = f"acc{book_id_number:05d}"

        # Update the record with the generated book_id
        cur.execute("""
            UPDATE books
            SET book_id = %s
            WHERE id = %s
        """, (book_id, book_id_number))

        mysql.connection.commit()
        cur.close()

        # Return the success message with Emalangeni
        return jsonify({
            "ok": True,
            "msg": f"Book added with ID: {book_id}, late fee per day: E{charge_after_due_date:.2f} (SZL)"
        })
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500


# --- Remove book by book_id (POST JSON/form) ---
@app.route('/books/remove', methods=['POST'])
def remove_book():
    data = request.get_json() or request.form
    book_id = data.get('book_id')
    if not book_id:
        return jsonify({"ok": False, "msg": "book_id required"}), 400
    try:
        cur = dict_cursor()
        cur.execute("SELECT id, status FROM books WHERE book_id=%s", (book_id,))
        row = cur.fetchone()
        if not row:
            cur.close()
            return jsonify({"ok": False, "msg": "Book not found"}), 404
        if row['status'] == 'Rented':
            cur.close()
            return jsonify({"ok": False, "msg": "Cannot remove a rented book"}), 400

        cur.execute("DELETE FROM books WHERE book_id=%s", (book_id,))
        mysql.connection.commit()
        cur.close()
        return jsonify({"ok": True, "msg": "Book removed"})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500


from datetime import datetime, date

@app.route('/books/checkout', methods=['POST'])
def checkout_book():
    data = request.get_json() or request.form

    member_id = data.get('member_id')
    if not member_id:
        return jsonify({"ok": False, "msg": "member_id is required"}), 400

    book_id = data.get('book_id')
    if not book_id:
        return jsonify({"ok": False, "msg": "book_id is required"}), 400

    due_date_str = data.get('due_date')
    if not due_date_str:
        return jsonify({"ok": False, "msg": "due_date is required"}), 400

    try:
        due_date = datetime.strptime(due_date_str, "%Y-%m-%d").date()
    except ValueError:
        return jsonify({"ok": False, "msg": "due_date must be in the format YYYY-MM-DD"}), 400

    today = date.today()
    # Validate that due date is not today or in the past
    if due_date <= today:
        # Option 1: Reject checkout with past/today due date
        # return jsonify({"ok": False, "msg": "Due date must be a future date"}), 400
        
        # Option 2: Allow but immediately mark as overdue
        overdue = True
    else:
        overdue = False

    try:
        cur = dict_cursor()

        # Check member existence and status
        cur.execute("SELECT status, full_name FROM members WHERE member_id=%s", (member_id,))
        member = cur.fetchone()
        if not member:
            return jsonify({"ok": False, "msg": f"Member with ID {member_id} not found"}), 404

        if member['status'] == 'suspended':
            return jsonify({"ok": False, "msg": "This member account is suspended. Cannot check out books."}), 403

        # Check if book exists and is available
        cur.execute("SELECT * FROM books WHERE book_id=%s", (book_id,))
        book = cur.fetchone()
        if not book:
            return jsonify({"ok": False, "msg": "Book not found"}), 404
        if book['status'] == 'Rented':
            return jsonify({"ok": False, "msg": "Book already rented"}), 400

        # Check outstanding balance
        cur.execute("""
            SELECT SUM(service_fee) AS outstanding_balance
            FROM books
            WHERE borrower_id = %s AND status = 'Rented'
        """, (member_id,))
        balance = cur.fetchone().get('outstanding_balance', 0) or 0
        if balance > 0:
            return jsonify({"ok": False, "msg": f"Member ID {member_id} currently owes E {balance}. Please settle before checkout."}), 403

        # Determine initial service fee if overdue
        initial_fee = 0.0
        if overdue:
            charge_after_due_date = book.get('charge_after_due_date', 0.0) or 0.0
            # Already overdue by 1 day
            initial_fee = round(charge_after_due_date, 2)

        borrower_name = member.get('full_name')
        cur.execute("""
            UPDATE books
            SET status='Rented',
                borrower_id=%s,
                borrower_name=%s,
                due_date=%s,
                service_fee=%s
            WHERE book_id=%s
        """, (member_id, borrower_name, due_date, initial_fee, book_id))

        
        if overdue and initial_fee > 0:
            cur.execute("""
                UPDATE members
                SET outstanding_fees = outstanding_fees + %s
                WHERE member_id=%s
            """, (initial_fee, member_id))

        mysql.connection.commit()
        return jsonify({
            "ok": True,
            "msg": f"Book '{book_id}' checked out to {borrower_name}" + (" (overdue!)" if overdue else ""),
            "service_fee": float(initial_fee)
        })

    except Exception as e:
        print(e)
        return jsonify({"ok": False, "msg": "Database error: " + str(e)}), 500
    finally:
        cur.close()


#return book route
@app.route('/books/return', methods=['POST'])
def return_book():
    data = request.get_json()
    if not data:
        return jsonify({"ok": False, "msg": "No data received"}), 400

    book_id = data.get('book_id')
    member_id = data.get('member_id')

    if not book_id or not member_id:
        return jsonify({"ok": False, "msg": "book_id and member_id are required"}), 400

    print(f"Received member_id: {member_id}")  # Debugging

    try:
        today = date.today()  # Use date for comparison
        cur = dict_cursor()

        cur.execute("SELECT * FROM books WHERE book_id=%s", (book_id,))
        book = cur.fetchone()
        if not book:
            cur.close()
            return jsonify({"ok": False, "msg": "Book not found"}), 404

        if book['status'] != 'Rented':
            cur.close()
            return jsonify({"ok": False, "msg": "Book is not currently rented"}), 400

        if book['borrower_id'] != member_id:
            cur.close()
            return jsonify({"ok": False, "msg": "This book was not rented by the specified member"}), 400

        due_date = book.get('due_date')
        fee = 0.0
        charge_after_due_date = book.get('charge_after_due_date', 0.0)

        if due_date:
            # Convert due_date to date if it's datetime
            if isinstance(due_date, datetime):
                due_date = due_date.date()

            if today > due_date:
                delta = (today - due_date).days
                fee = round(delta * charge_after_due_date, 2)

        # Update book status
        cur.execute("""
            UPDATE books
            SET status='Available',
                borrower_id=NULL,
                borrower_name=NULL,
                due_date=NULL,
                service_fee=%s
            WHERE book_id=%s
        """, (fee, book_id))

        # Update member's outstanding fees
        cur.execute("UPDATE members SET outstanding_fees = outstanding_fees + %s WHERE member_id = %s", (fee, member_id))

        mysql.connection.commit()
        cur.close()

        return jsonify({"ok": True, "msg": "Book returned", "service_fee": float(fee)})

    except Exception as e:
        print(f"Error occurred: {e}")
        return jsonify({"ok": False, "msg": str(e)}), 500


@app.route('/members/list')
def list_members():
    cur = mysql.connection.cursor()
    cur.execute("SELECT member_id, full_name, email FROM members ORDER BY member_id")
    members = cur.fetchall()
    return jsonify({'members': members})

# Remove member
@app.route('/members/remove', methods=['POST'])
def remove_member():
    data = request.json
    member_id = data.get('member_id')
    if not member_id:
        return jsonify({'ok': False, 'msg': 'Member ID required'}), 400
    cur = mysql.connection.cursor()
    try:
        cur.execute("DELETE FROM members WHERE member_id = %s", (member_id,))
        mysql.connection.commit()
        return jsonify({'ok': True, 'msg': 'Member removed'})
    except Exception as e:
        return jsonify({'ok': False, 'msg': str(e)}), 500

# Reports

from datetime import date

from datetime import date

@app.route('/reports')
def reports_page():
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # --- Count totals ---
    cur.execute("SELECT COUNT(*) AS cnt FROM books")
    total_books = cur.fetchone().get('cnt', 0)

    cur.execute("SELECT COUNT(*) AS cnt FROM books WHERE status = 'Rented'")
    rented_books = cur.fetchone().get('cnt', 0)

    cur.execute("SELECT COUNT(*) AS cnt FROM books WHERE status = 'Rented' AND due_date < CURDATE()")
    overdue_books = cur.fetchone().get('cnt', 0)

   
    cur.execute("""
        SELECT
            b.book_id,
            b.title,
            COALESCE(b.borrower_name, m.full_name) AS borrower_name,
            DATE_FORMAT(b.due_date, '%Y-%m-%d') AS due_date,
            b.charge_after_due_date
        FROM books b
        LEFT JOIN members m ON b.member_id = m.member_id
        WHERE b.status = 'Rented' AND b.due_date < CURDATE()
        ORDER BY b.due_date ASC
    """)
    overdue_list_raw = cur.fetchall()

    # Calculate due_amount
    today = date.today()
    overdue_list = []
    for b in overdue_list_raw:
        due_date_obj = date.fromisoformat(b['due_date'])
        days_overdue = (today - due_date_obj).days
        charge = b.get('charge_after_due_date') or 0
        b['due_amount'] = days_overdue * charge
        overdue_list.append(b)
    #cur.execute("update members set outstanding_fees=%s where member_id=%s")
    # --- Total members ---
    cur.execute("SELECT COUNT(*) AS cnt FROM members")
    total_members = cur.fetchone().get('cnt', 0)

    # --- Suspended members ---
    cur.execute("SELECT COUNT(*) AS cnt FROM members WHERE status='suspended'")
    suspended_members = cur.fetchone().get('cnt', 0)

    #Members by age group
    cur.execute("""
        SELECT
            age_group AS age_group,
            COUNT(*) AS cnt
        FROM members
        WHERE age_group IS NOT NULL AND TRIM(age_group) <> ''
        GROUP BY age_group
        ORDER BY FIELD(age_group, '0-18','19-35','36-50','51+')
    """)
    age_ranges = cur.fetchall()

    # Checked out books 
    cur.execute("""
        SELECT
            b.book_id,
            b.title,
            COALESCE(b.borrower_name, m.full_name) AS borrower_name,
            DATE_FORMAT(b.due_date, '%Y-%m-%d') AS due_date
        FROM books b
        LEFT JOIN members m ON b.member_id = m.member_id
        WHERE b.status = 'Rented'
        ORDER BY b.due_date ASC
    """)
    checked_out_books = cur.fetchall()
    checked_out_count = len(checked_out_books)  # Count for card

    cur.close()

    report = {
        'total_books': total_books,
        'rented_books': rented_books,
        'checked_out_count': checked_out_count,
        'overdue_books': overdue_books,
        'overdue_list': overdue_list,
        'total_members': total_members,
        'suspended_members': suspended_members,
        'age_ranges': age_ranges,
        'checked_out_books': checked_out_books
    }

    return render_template('reports.html', report=report, username=session.get("username"))



@app.route('/generate_report')
def generate_report():
    # Logic to generate a new report goes here
    flash("Report generated successfully!", "success")
    return redirect(url_for('reports_page'))


@app.route('/export_report')
def export_report():
    cur = mysql.connection.cursor()

    # 1. Summary / Monthly Stats 
    last_month = datetime.now() - timedelta(days=30)

    # Total members
    cur.execute("SELECT COUNT(*) FROM members")
    total_members = cur.fetchone()[0]

    # Suspended members
    cur.execute("SELECT COUNT(*) FROM members WHERE status = 'suspended'")
    suspended_members = cur.fetchone()[0]

    # Total books
    cur.execute("SELECT COUNT(*) FROM books")
    total_books = cur.fetchone()[0]

    # Books currently rented
    cur.execute("SELECT COUNT(*) FROM books WHERE status = 'Rented'")
    rented_count = cur.fetchone()[0]

    # New books added last month
    cur.execute("SELECT COUNT(*) FROM books WHERE created_at >= %s", (last_month,))
    new_books_last_month = cur.fetchone()[0]

    # New members last month
    cur.execute("SELECT COUNT(*) FROM members WHERE join_date >= %s", (last_month,))
    new_members_last_month = cur.fetchone()[0]

    summary_data = [{
        'Total Members': total_members,
        'Suspended Members': suspended_members,
        'Total Books': total_books,
        'Books Rented': rented_count,
        'New Books (Last 30 days)': new_books_last_month,
        'New Members (Last 30 days)': new_members_last_month
    }]
    df_summary = pd.DataFrame(summary_data)

    # 2. Rented Books 
    cur.execute("SELECT book_id, title, borrower_name, due_date FROM books WHERE status = 'Rented'")
    rented_books = cur.fetchall()
    df_rented = pd.DataFrame(rented_books, columns=['Book ID', 'Title', 'Borrower', 'Due Date'])

    # 3. Overdue Books 
    cur.execute("""
        SELECT book_id, title, borrower_name, due_date, DATEDIFF(CURDATE(), due_date) AS Days_Overdue
        FROM books
        WHERE status = 'Rented' AND due_date < CURDATE()
    """)
    overdue_books = cur.fetchall()
    df_overdue = pd.DataFrame(overdue_books, columns=['Book ID', 'Title', 'Borrower', 'Due Date', 'Days Overdue'])

    # --- 4. Newly Added Books (Last Month) ---
    cur.execute("SELECT book_id, title, author, category, created_at FROM books WHERE created_at >= %s", (last_month,))
    new_books = cur.fetchall()
    df_new_books = pd.DataFrame(new_books, columns=['Book ID', 'Title', 'Author', 'Category', 'Added Date'])

    # --- 5. Members ---
    cur.execute("SELECT member_id, full_name, email, phone, join_date FROM members")
    members = cur.fetchall()
    df_members = pd.DataFrame(members, columns=['Member ID', 'Name', 'Email', 'Phone', 'Date Joined'])

    # --- 6. Suspended Members ---
    cur.execute("SELECT member_id, full_name, email, phone, join_date FROM members WHERE status = 'suspended'")
    suspended = cur.fetchall()
    df_suspended = pd.DataFrame(suspended, columns=['Member ID', 'Name', 'Email', 'Phone', 'Date Joined'])

    # --- 7. Book Inventory ---
    cur.execute("SELECT book_id, title, author, category, status, borrower_name FROM books")
    books = cur.fetchall()
    df_books = pd.DataFrame(books, columns=['Book ID', 'Title', 'Author', 'Category', 'Status', 'Borrower'])

    cur.close()

    # Create Excel file in memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df_summary.to_excel(writer, index=False, sheet_name='Summary')
        df_rented.to_excel(writer, index=False, sheet_name='Rented Books')
        df_overdue.to_excel(writer, index=False, sheet_name='Overdue Books')
        df_new_books.to_excel(writer, index=False, sheet_name='New Books')
        df_members.to_excel(writer, index=False, sheet_name='Members')
        df_suspended.to_excel(writer, index=False, sheet_name='Suspended Members')
        df_books.to_excel(writer, index=False, sheet_name='Book Inventory')

    output.seek(0)

    flash("Report exported successfully!", "success")
    return send_file(output,
                     as_attachment=True,
                     download_name='library_report.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')



@app.route('/view_statistics')
def view_statistics():
    if "username" not in session:
        flash("Please log in first.", "danger")
        return redirect(url_for("login"))

    
    cur = dict_cursor()

    # 1. Total members
  
    cur.execute("SELECT COUNT(*) AS total FROM members")
    result = cur.fetchone()
    total_members = result['total'] if result else 0

    
    # 2. Age group statistics
    
    cur.execute("""
        SELECT age_group, COUNT(*) AS count
        FROM members
        GROUP BY age_group
    """)
    age_statistics = cur.fetchall() or []

    age_labels = [row['age_group'] for row in age_statistics]
    age_counts = [row['count'] for row in age_statistics]

   
    # 3. Book status statistics
    
    cur.execute("""
        SELECT status, COUNT(*) AS count
        FROM books
        GROUP BY status
    """)
    book_statistics = cur.fetchall() or []

    book_labels = [row['status'] for row in book_statistics]
    book_counts = [row['count'] for row in book_statistics]

    cur.close()

    return render_template(
        'statistics.html',
        total_members=total_members,
        age_labels=json.dumps(age_labels),
        age_counts=json.dumps(age_counts),
        book_labels=json.dumps(book_labels),
        book_counts=json.dumps(book_counts),
        username=session.get("username")
    )

# Route to serve the registration page
@app.route('/members/register', methods=['GET'])
def member_registration():
    return render_template('member_registration.html')

# Route to serve the registration page
from datetime import datetime
from zoneinfo import ZoneInfo

def _parse_test_date(s: str):
    
    if not s:
        return None
    s = s.strip()
    patterns = ["%Y-%m-%d", "%Y%m%d", "%Y-%m", "%Y%m"]
    for p in patterns:
        try:
            return datetime.strptime(s, p).date()
        except Exception:
            continue
    # last attempt: ISO parse (allow YYYY-MM-DDTHH:MM)
    try:
        return datetime.fromisoformat(s).date()
    except Exception:
        return None

@app.route('/members/add', methods=['POST'])
def add_member():
    import traceback
    data = request.get_json() or {}

    # --- Extract and validate input fields (kept from your original) ---
    gender = data.get('gender')
    if isinstance(gender, str):
        gender = gender[:10]
    else:
        gender = None

    full_name = data.get('full_name') or "Unknown Member"
    email = data.get('email')
    phone = data.get('phone')
    age = data.get('age')
    physical_address = data.get('physical_address')
    postal_address = data.get('postal_address')
    join_date = data.get('join_date') or datetime.today().isoformat()
    status = data.get('status') or 'active'
    outstanding_fees = data.get('outstanding_fees') if data.get('outstanding_fees') is not None else 0.0
    notes = data.get('notes')
    next_of_kin = data.get('next_of_kin')
    next_of_kin_contact = data.get('next_of_kin_contact')
    age_group = data.get('age_group')

    # Testing helpers
    test_date_raw = data.get('test_date')           
    dry_run = bool(data.get('dry_run'))             
    test_date = _parse_test_date(test_date_raw)     

    # --- Handle member_id ---
    input_member_id = data.get('member_id')
    member_id = None

    try:
        recconnect_mysql()
        cur = dict_cursor()

        # If caller supplied a member_id, respect it (no change)
        if input_member_id:
            member_id_str = str(input_member_id)
            if member_id_str.startswith('M') or member_id_str.isdigit():
                member_id = member_id_str if member_id_str.startswith('M') else f"M{member_id_str}"
            else:
                member_id = member_id_str

            # Insert with supplied id (works as before)
            cur.execute("""
                INSERT INTO members
                (member_id, gender, full_name, email, phone, physical_address, postal_address, join_date,
                 status, outstanding_fees, notes, age, next_of_kin, next_of_kin_contact, age_group)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                member_id, gender, full_name, email, phone, physical_address, postal_address, join_date,
                status, outstanding_fees, notes, age, next_of_kin, next_of_kin_contact, age_group
            ))
            mysql.connection.commit()
            cur.close()
            return jsonify({"ok": True, "msg": "Inserted with provided member_id", "member_id": member_id}), 201

        # Determine which date to use (test_date overrides real now)
        TZ = ZoneInfo("Africa/Mbabane")
        if test_date:
            # create a datetime in timezone for consistent formatting
            now = datetime(test_date.year, test_date.month, test_date.day, tzinfo=TZ)
        else:
            now = datetime.now(TZ)

        date_prefix = now.strftime("%y%m")   # e.g. "2509"

        # Suffix config (keeps your BASE suffix behavior)
        SUFFIX_WIDTH = 4
        BASE_SUFFIX = 794   # numeric form of 0794

        
        if dry_run:
          
            cur.execute("SELECT last_suffix FROM member_suffix_counter WHERE id = 1;")
            row = cur.fetchone()
            last_suffix = int(row['last_suffix']) if (row and row.get('last_suffix') is not None) else (BASE_SUFFIX - 1)
            new_suffix = last_suffix + 1
            suffix_str = str(new_suffix).zfill(SUFFIX_WIDTH)
            candidate_member_id = f"M{date_prefix}{suffix_str}"
            cur.close()
            return jsonify({"ok": True, "dry_run": True, "member_id": candidate_member_id, "date_prefix": date_prefix, "suffix": suffix_str}), 200

        
        cur.execute("""
            CREATE TABLE IF NOT EXISTS member_suffix_counter (
                id TINYINT PRIMARY KEY,
                last_suffix INT NOT NULL
            ) ENGINE=InnoDB;
        """)

        # Transaction & locking
        cur.execute("START TRANSACTION;")
        cur.execute("SELECT last_suffix FROM member_suffix_counter WHERE id = 1 FOR UPDATE;")
        row = cur.fetchone()
        if row and row.get('last_suffix') is not None:
            last_suffix = int(row['last_suffix'])
        else:
            last_suffix = BASE_SUFFIX - 1
            # insert initial row (if missing)
            cur.execute("INSERT INTO member_suffix_counter (id, last_suffix) VALUES (1, %s) ON DUPLICATE KEY UPDATE last_suffix = last_suffix;", (last_suffix,))

        new_suffix = last_suffix + 1
        cur.execute("UPDATE member_suffix_counter SET last_suffix = %s WHERE id = 1;", (new_suffix,))

        suffix_str = str(new_suffix).zfill(SUFFIX_WIDTH)
        candidate_member_id = f"M{date_prefix}{suffix_str}"   # e.g. "25090794"

        # collision check
        cur.execute("SELECT 1 FROM members WHERE member_id = %s LIMIT 1;", (candidate_member_id,))
        if cur.fetchone():
            # collision: rollback and try once more (simple handling)
            cur.execute("ROLLBACK;")
            # increment once more and try again
            cur.execute("START TRANSACTION;")
            cur.execute("SELECT last_suffix FROM member_suffix_counter WHERE id = 1 FOR UPDATE;")
            row2 = cur.fetchone()
            last_suffix2 = int(row2['last_suffix'])
            new_suffix2 = last_suffix2 + 1
            cur.execute("UPDATE member_suffix_counter SET last_suffix = %s WHERE id = 1;", (new_suffix2,))
            suffix_str = str(new_suffix2).zfill(SUFFIX_WIDTH)
            candidate_member_id = f"M{date_prefix}{suffix_str}"

        # finally insert member with generated id
        cur.execute("""
            INSERT INTO members
            (member_id, gender, full_name, email, phone, physical_address, postal_address, join_date,
             status, outstanding_fees, notes, age, next_of_kin, next_of_kin_contact, age_group)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            candidate_member_id, gender, full_name, email, phone, physical_address, postal_address, join_date,
            status, outstanding_fees, notes, age, next_of_kin, next_of_kin_contact, age_group
        ))

        mysql.connection.commit()
        cur.close()
        return jsonify({"ok": True, "msg": "Member registered successfully.", "member_id": candidate_member_id}), 201

    except Exception as e:
        print("Error occurred during member insert:")
        traceback.print_exc()
        try:
            cur.execute("ROLLBACK;")
        except Exception:
            pass
        return jsonify({"ok": False, "msg": "Database error."}), 500

# ---- View Members Page -------------------------------------
@app.route('/members/view', methods=['GET'])
def view_members():
    if "username" not in session:
        flash("Please log in first.", "danger")
        return redirect(url_for("login"))
    return render_template('view_members.html')

# ---- API: List Members ----
@app.route('/api/members', methods=['GET'])
def api_list_members():
    try:
        cur = dict_cursor()
        cur.execute("""
            SELECT
              member_id, gender, full_name, email, phone,
              physical_address, postal_address, join_date, status,
              outstanding_fees, notes, age, next_of_kin, next_of_kin_contact
            FROM members
            ORDER BY member_id
        """)
        members = cur.fetchall()
        cur.close()
        return jsonify({'ok': True, 'members': members})
    except Exception as e:
        return jsonify({'ok': False, 'msg': f"Error fetching members: {e}"}), 500

# ---- API: Delete a Member ----
@app.route('/api/members/delete', methods=['POST'])
def api_delete_member():
    data = request.get_json() or {}
    member_id = data.get('member_id')

    if not member_id:
        return jsonify({'ok': False, 'msg': 'member_id is required'}), 400

    try:
        cur = dict_cursor()
        cur.execute("DELETE FROM members WHERE member_id = %s", (member_id,))
        mysql.connection.commit()
        cur.close()
        return jsonify({'ok': True, 'msg': f'Member {member_id} deleted'})
    except Exception as e:
        return jsonify({'ok': False, 'msg': f"Error deleting member: {e}"}), 500

# ---- API: Update Member Status ----
# ---- API: Update Member Status ----
@app.route('/api/members/update_status', methods=['POST'])
def api_update_member_status():
    data = request.get_json() or {}
    member_id = data.get('member_id')
    status = data.get('status')

    if not member_id or not status:
        return jsonify({'ok': False, 'msg': 'member_id and status are required'}), 400

    valid_statuses = ('active', 'inactive', 'suspended')
    if status.lower() not in valid_statuses:
        return jsonify({'ok': False, 'msg': f'Invalid status. Must be one of {valid_statuses}'}), 400

    try:
        cur = dict_cursor()
        cur.execute("UPDATE members SET status = %s WHERE member_id = %s", (status.lower(), member_id))
        mysql.connection.commit()
        cur.close()
        return jsonify({'ok': True, 'msg': f'Status updated to {status.lower()} for member {member_id}'})
    except Exception as e:
        print(e)  # error logging
        return jsonify({'ok': False, 'msg': f"Error updating status: {e}"}), 500


@app.route('/daily_planner', methods=['GET', 'POST'])
def daily_planner():
    recconnect_mysql()  

    if request.method == 'POST':
        task_title = request.form['task-title']
        task_date = request.form['task-date']
        task_notes = request.form['task-notes']

        # Insert the new task into the database
        cur = dict_cursor()
        cur.execute(
            'INSERT INTO tasks (title, due_date, notes) VALUES (%s, %s, %s)',
            (task_title, task_date, task_notes)
        )
        mysql.connection.commit()
        cur.close()

        flash('Task added successfully!', 'success')  # Add a success message
        return redirect(url_for('daily_planner'))

    # Retrieve tasks from the database
    cur = dict_cursor()
    cur.execute('SELECT * FROM tasks ORDER BY due_date')
    tasks = cur.fetchall()
    cur.close()

    return render_template('daily_planner.html', tasks=tasks)

@app.route('/api/view_member')
def api_view_member():
    member_id = request.args.get('member_id')
    if not member_id:
        return jsonify({'error': 'Member ID is required'}), 400

    cur = dict_cursor()

    try:
        cur.execute("SELECT * FROM members WHERE member_id = %s", (member_id,))
        member = cur.fetchone()
    except Exception as e:
        cur.close()
        return jsonify({'error': f'Database error: {str(e)}'}), 500

    cur.close()

    if member:
        # Directly return the full dictionary from DictCursor
        return jsonify(member)
    else:
        return jsonify({'error': 'Member not found'}), 404


# Define the upload folder
UPLOAD_FOLDER = 'uploads/'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

from datetime import datetime

@app.route('/api/import_members', methods=['POST'])
def import_members():
    if 'file' not in request.files:
        return jsonify({'ok': False, 'msg': 'No file part'}), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({'ok': False, 'msg': 'No selected file'}), 400

    # Check file extension
    allowed_extensions = {'.csv', '.xlsx'}
    file_extension = os.path.splitext(file.filename)[1].lower()

    if file_extension not in allowed_extensions:
        return jsonify({'ok': False, 'msg': 'Only CSV and XLSX files are allowed'}), 400

    # Save the file
    filepath = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(filepath)

    imported_members = []
    skipped_members = []   # will hold dicts with reason

    try:
        # We'll need DB access to check duplicates
        recconnect_mysql()
        cur = dict_cursor()

        # helper to test duplicate existence
        def is_duplicate(member_data):
            """
            Check if a member already exists.
            Priority: member_id (if provided) -> email -> phone
            Returns (True, reason) or (False, None).
            """
            mid = member_data.get('member_id')
            email = member_data.get('email')
            phone = member_data.get('phone')

            if mid:
                # normalize: strip leading/trailing whitespace
                mm = str(mid).strip()
                cur.execute("SELECT 1 FROM members WHERE member_id = %s LIMIT 1;", (mm,))
                if cur.fetchone():
                    return True, f"member_id {mm} exists"

            if email:
                ee = str(email).strip().lower()
                cur.execute("SELECT 1 FROM members WHERE email = %s LIMIT 1;", (ee,))
                if cur.fetchone():
                    return True, f"email {ee} exists"

            if phone:
                pp = str(phone).strip()
                cur.execute("SELECT 1 FROM members WHERE phone = %s LIMIT 1;", (pp,))
                if cur.fetchone():
                    return True, f"phone {pp} exists"

            return False, None

        # CSV path
        if file_extension == '.csv':
            empty_row_streak = 0
            with open(filepath, newline='', encoding='utf-8-sig') as csvfile:
                reader = csv.DictReader(csvfile)
                for raw_row in reader:
                    # Normalize values: strip strings, treat empty strings as None
                    row = {k: (v.strip() if isinstance(v, str) else v) for k, v in raw_row.items()}

                    # If row entirely empty (no values or only blanks)
                    if all(not v for v in row.values()):
                        empty_row_streak += 1
                        if empty_row_streak >= 3:
                            # stop importing after 3 consecutive empty rows
                            print("📂 Detected 3 consecutive empty rows — stopping import.")
                            break
                        else:
                            # skip this empty row but continue reading
                            continue
                    else:
                        empty_row_streak = 0  # reset streak

                    member_data = prepare_member_data(row)

                    # check duplicates
                    dup, reason = is_duplicate(member_data)
                    if dup:
                        skipped_members.append({'member_data': member_data, 'reason': reason})
                        continue

                    # call your add endpoint
                    resp = requests.post('http://localhost:5000/members/add', json=member_data, timeout=30)
                    if resp.ok:
                        imported_members.append(resp.json())
                    else:
                        # if add failed because of duplicate or other DB error, record and continue
                        try:
                            msg = resp.json().get('msg')
                        except Exception:
                            msg = resp.text
                        skipped_members.append({'member_data': member_data, 'reason': f'add_failed: {msg}'})
                        # continue importing other rows

        # XLSX path
        elif file_extension == '.xlsx':
            wb = openpyxl.load_workbook(filepath, read_only=True)
            sheet = wb.active
            empty_row_streak = 0
            # assume header is in row 1; iterate from row 2
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # treat None or empty-string cells as empty
                if all((cell is None or (isinstance(cell, str) and cell.strip() == "")) for cell in row):
                    empty_row_streak += 1
                    if empty_row_streak >= 3:
                        print("📂 Detected 3 consecutive empty rows — stopping import.")
                        break
                    else:
                        continue
                else:
                    empty_row_streak = 0

                # map columns into keys (adjust indices if your sheet layout differs)
                # Here: col0=full_name, col1=gender, col2=member_id, etc.
                row_dict = {
                    'full_name': row[0] if len(row) > 0 else None,
                    'gender': row[1] if len(row) > 1 else None,
                    'member_id': row[2] if len(row) > 2 else None,
                    'email': row[3] if len(row) > 3 else None,
                    'phone': row[4] if len(row) > 4 else None,
                    # add more mappings if needed
                }

                # normalize strings
                row_dict = {k: (v.strip() if isinstance(v, str) else v) for k, v in row_dict.items()}

                member_data = {
                    'full_name': row_dict.get('full_name'),
                    'gender': row_dict.get('gender'),
                    'member_id': row_dict.get('member_id'),
                    
                }

                dup, reason = is_duplicate(member_data)
                if dup:
                    skipped_members.append({'member_data': member_data, 'reason': reason})
                    continue
                print("📘 Inserting book (CSV):", member_data)
                resp = requests.post('http://localhost:5000/members/add', json=member_data, timeout=30)
                if resp.ok:
                    imported_members.append(resp.json())
                else:
                    try:
                        msg = resp.json().get('msg')
                    except Exception:
                        msg = resp.text
                    skipped_members.append({'member_data': member_data, 'reason': f'add_failed: {msg}'})

        # done reading file
        cur.close()
        return jsonify({
            'ok': True,
            'msg': 'Members import completed',
            'imported_count': len(imported_members),
            'skipped_count': len(skipped_members),
            'imported': imported_members,
            'skipped': skipped_members
        }), 200

    except Exception as e:
        try:
            cur.close()
        except Exception:
            pass
        return jsonify({'ok': False, 'msg': f'Error processing file: {str(e)}'}), 500


def prepare_member_data(row):
    # Prepare member data from CSV row dict (keys expected from CSV header)
    # Normalizes values (strip strings) and returns a dict suitable for /members/add
    def norm(v):
        if v is None:
            return None
        if isinstance(v, str):
            s = v.strip()
            return s if s != "" else None
        return v

    return {
        'full_name': norm(row.get('full_name')),
        'gender': norm(row.get('gender')),
        'member_id': norm(row.get('member_id')),
        'username': None,
        'email': norm(row.get('email')),
        'phone': norm(row.get('phone')),
        'physical_address': norm(row.get('physical_address')),
        'postal_address': norm(row.get('postal_address')),
        'join_date': norm(row.get('join_date')),
        'status': norm(row.get('status')),
        'outstanding_fees': norm(row.get('outstanding_fees')),
        'notes': norm(row.get('notes')),
        'age': norm(row.get('age')),
        'next_of_kin': norm(row.get('next_of_kin')),
        'next_of_kin_contact': norm(row.get('next_of_kin_contact')),
        'age_group': norm(row.get('age_group')),
    }



import traceback

ALLOWED_EXT = {'.csv', '.xlsx'}
@app.route('/api/import_books', methods=['POST'])
def import_books():
    import traceback
    if 'file' not in request.files:
        return jsonify({'ok': False, 'msg': 'No file part'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'ok': False, 'msg': 'No selected file'}), 400

    file_extension = os.path.splitext(file.filename)[1].lower()
    if file_extension not in ALLOWED_EXT:
        return jsonify({'ok': False, 'msg': 'Only CSV and XLSX files are allowed'}), 400

    filepath = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(filepath)

    imported_books = []
    skipped_books = []  # will hold dicts with reason

    try:
        recconnect_mysql()
        cur = dict_cursor()

        # helper to test duplicate existence
        def is_duplicate(book_data):
            """
            Check if a book already exists by book_id or title+author.
            Returns (True, reason) or (False, None)
            """
            bid = book_data.get('book_id')
            title = book_data.get('title')
            author = book_data.get('author')

            if bid:
                cur.execute("SELECT 1 FROM books WHERE book_id = %s LIMIT 1;", (bid,))
                if cur.fetchone():
                    return True, f"book_id {bid} exists"

            if title and author:
                cur.execute("SELECT 1 FROM books WHERE title = %s AND author = %s LIMIT 1;", (title, author))
                if cur.fetchone():
                    return True, f"title+author {title} / {author} exists"

            return False, None

        # --- CSV processing ---
        if file_extension == '.csv':
            empty_row_streak = 0
            with open(filepath, newline='', encoding='utf-8-sig') as csvfile:
                reader = csv.DictReader(csvfile)
                for row in reader:
                    if all(not value for value in row.values()):
                        empty_row_streak += 1
                        if empty_row_streak >= 3:
                            print("📂 Detected 3 consecutive empty rows — stopping import.")
                            break
                        continue
                    else:
                        empty_row_streak = 0

                    book_data = prepare_book_data(row)

                    # Ensure year is an int
                    try:
                        book_data['year'] = int(book_data['year'])
                    except (TypeError, ValueError):
                        book_data['year'] = 0

                    # Ensure required fields
                    book_data['title'] = book_data.get('title') or '-'
                    book_data['category'] = book_data.get('category') or '-'
                    book_data['status'] = book_data.get('status') if book_data.get('status') in ('Available', 'Rented') else 'Available'

                    # check duplicates
                    dup, reason = is_duplicate(book_data)
                    if dup:
                        skipped_books.append({'book_data': book_data, 'reason': reason})
                        continue

                    print("📘 Inserting book (CSV):", book_data)
                    response = requests.post('http://localhost:5000/books/add', json=book_data)
                    if response.ok:
                        try:
                            imported_books.append(response.json())
                        except Exception:
                            imported_books.append({'ok': True, 'msg': 'Added (no JSON)', 'data': book_data})
                    else:
                        try:
                            err = response.json().get('msg')
                        except Exception:
                            err = response.text or f'status {response.status_code}'
                        skipped_books.append({'book_data': book_data, 'reason': f'add_failed: {err}'})
                        continue

        # --- XLSX processing ---
        elif file_extension == '.xlsx':
            workbook = openpyxl.load_workbook(filepath, data_only=True)
            sheet = workbook.active
            empty_row_streak = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if all(cell is None for cell in row):
                    empty_row_streak += 1
                    if empty_row_streak >= 3:
                        print("📂 Detected 3 consecutive empty rows — stopping import.")
                        break
                    continue
                else:
                    empty_row_streak = 0

                year_val = row[4] if len(row) > 4 else None
                try:
                    year_val = int(year_val)
                except (TypeError, ValueError):
                    year_val = 0

                book_data = {
                    'book_id': row[0] if len(row) > 0 else None,
                    'title': row[1] if len(row) > 1 and row[1] else '-',
                    'author': row[2] if len(row) > 2 and row[2] else '-',
                    'publisher': row[3] if len(row) > 3 and row[3] else '-',
                    'year': year_val,
                    'category': row[5] if len(row) > 5 and row[5] else '-',
                    'status': row[6] if len(row) > 6 and row[6] in ('Available', 'Rented') else 'Available'
                }

                # check duplicates
                dup, reason = is_duplicate(book_data)
                if dup:
                    skipped_books.append({'book_data': book_data, 'reason': reason})
                    continue

                print("📘 Inserting book (XLSX):", book_data)
                response = requests.post('http://localhost:5000/books/add', json=book_data)
                if response.ok:
                    try:
                        imported_books.append(response.json())
                    except Exception:
                        imported_books.append({'ok': True, 'msg': 'Added (no JSON)', 'data': book_data})
                else:
                    try:
                        err = response.json().get('msg')
                    except Exception:
                        err = response.text or f'status {response.status_code}'
                    skipped_books.append({'book_data': book_data, 'reason': f'add_failed: {err}'})
                    continue

        # Remove uploaded file after processing
        try:
            os.remove(filepath)
        except Exception:
            pass

        return jsonify({
            'ok': True,
            'msg': 'Books import completed',
            'imported_count': len(imported_books),
            'skipped_count': len(skipped_books),
            'imported': imported_books,
            'skipped': skipped_books
        }), 200

    except Exception as e:
        traceback.print_exc()
        try:
            if os.path.exists(filepath):
                os.remove(filepath)
        except Exception:
            pass
        return jsonify({'ok': False, 'msg': f'Error processing file: {str(e)}'}), 500


if __name__ == "__main__":
    app.run(host='0.0.0.0', debug=True, port=5000)





